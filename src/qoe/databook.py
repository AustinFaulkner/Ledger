"""Excel "databook" export — the supporting workbook a QoE analyst hands over alongside
the PDF report. One sheet per schedule: Summary, Financials, EBITDA Bridge, Working
Capital, Red Flags, Readiness, Data Profile. Pulls the cached AI results plus the offline
financial series / readiness / profiling. openpyxl only (already a dependency).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from . import financials, profiling, readiness, store

TITLE = Font(bold=True, size=14)
BOLD = Font(bold=True)


def build(project_id: str) -> bytes:
    """Assemble the databook workbook and return it as .xlsx bytes."""
    proj = store.get_project(project_id) or {}
    fin = financials.extract(project_id)
    qoe = store.get_report(project_id, "qoe")
    rf = store.get_report(project_id, "redflags")
    wc = store.get_report(project_id, "working_capital")
    rd = readiness.assess(project_id)
    profs = profiling.profile_project(project_id)

    wb = Workbook()
    _summary(wb.active, proj, fin, qoe, rf, wc, rd)
    _financials(wb.create_sheet("Financials"), fin)
    if qoe:
        _bridge(wb.create_sheet("EBITDA Bridge"), qoe["payload"])
    if wc:
        _working_capital(wb.create_sheet("Working Capital"), wc["payload"])
    if rf:
        _redflags(wb.create_sheet("Red Flags"), rf["payload"])
    _readiness(wb.create_sheet("Readiness"), rd)
    if profs:
        _profile(wb.create_sheet("Data Profile"), profs)

    for ws in wb.worksheets:
        _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _last(vals):
    if not vals:
        return None
    for v in reversed(vals):
        if v is not None:
            return v
    return None


def _summary(ws, proj, fin, qoe, rf, wc, rd):
    ws.title = "Summary"
    ws.cell(1, 1, proj.get("company") or proj.get("name") or "Deal").font = TITLE
    ws.cell(2, 1, "Quality of Earnings & Diligence — Databook ($ in thousands)")
    metrics = [
        ("Revenue (latest)", _last(fin["series"].get("Revenue"))),
        ("EBITDA (latest)", _last(fin["series"].get("EBITDA"))),
        ("Normalized EBITDA", qoe["payload"]["normalized_ebitda"] if qoe else None),
        ("Suggested NWC peg", wc["payload"].get("suggested_peg") if wc else None),
        ("Red flags", len(rf["payload"]["findings"]) if rf else None),
        ("Data-room readiness", f"{rd['score']}%"),
    ]
    r = 4
    for label, val in metrics:
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 2, val if val is not None else "—")
        r += 1
    present = [n for n, ok in [("QoE", bool(qoe)), ("Red Flags", bool(rf)), ("Working Capital", bool(wc))] if ok]
    ws.cell(r + 1, 1, "AI sections present").font = BOLD
    ws.cell(r + 1, 2, ", ".join(present) or "none yet — run the analyses to populate")


def _financials(ws, fin):
    ws.cell(1, 1, "Financials ($000s)").font = TITLE
    periods = fin["periods"]
    r = 3
    ws.cell(r, 1, "Line item").font = BOLD
    for j, p in enumerate(periods):
        ws.cell(r, 2 + j, p).font = BOLD
    r += 1
    for name, vals in fin["series"].items():
        ws.cell(r, 1, name)
        for j, v in enumerate(vals):
            ws.cell(r, 2 + j, v)
        r += 1
    if fin["margins"]:
        r += 1
        ws.cell(r, 1, "Margins (%)").font = BOLD
        r += 1
        for name, vals in fin["margins"].items():
            ws.cell(r, 1, name)
            for j, v in enumerate(vals):
                ws.cell(r, 2 + j, v)
            r += 1


def _bridge(ws, q):
    ws.cell(1, 1, "EBITDA Bridge ($000s)").font = TITLE
    ws.cell(3, 1, "Reported EBITDA").font = BOLD
    ws.cell(3, 2, q.get("reported_ebitda"))
    r = 5
    for j, h in enumerate(["Adjustment", "Amount", "Type", "Confidence", "Rationale", "Evidence source", "Evidence quote"]):
        ws.cell(r, 1 + j, h).font = BOLD
    r += 1
    for a in q.get("adjustments", []):
        ws.cell(r, 1, a.get("label"))
        ws.cell(r, 2, a.get("amount"))
        ws.cell(r, 3, (a.get("kind") or "").replace("_", " "))
        ws.cell(r, 4, a.get("confidence"))
        ws.cell(r, 5, a.get("rationale"))
        ws.cell(r, 6, a.get("evidence_source"))
        ws.cell(r, 7, a.get("evidence_quote"))
        r += 1
    ws.cell(r + 1, 1, "Normalized EBITDA").font = BOLD
    ws.cell(r + 1, 2, q.get("normalized_ebitda"))


def _working_capital(ws, w):
    ws.cell(1, 1, "Working Capital & Proof of Cash ($000s)").font = TITLE
    state = {"r": 3}

    def table(title, items):
        r = state["r"]
        ws.cell(r, 1, title).font = BOLD
        r += 1
        for j, h in enumerate(["Line item", "Amount", "Evidence source"]):
            ws.cell(r, 1 + j, h).font = BOLD
        r += 1
        for li in items:
            ws.cell(r, 1, li.get("label"))
            ws.cell(r, 2, li.get("amount"))
            ws.cell(r, 3, li.get("evidence_source"))
            r += 1
        state["r"] = r + 1

    table("Current assets", w.get("current_assets", []))
    table("Current liabilities", w.get("current_liabilities", []))
    if w.get("debt_like_items"):
        table("Debt-like items (EV → equity)", w["debt_like_items"])

    r = state["r"]
    for label, key in [("Net working capital", "net_working_capital"), ("NWC % of revenue", "nwc_pct_of_revenue"), ("Suggested peg", "suggested_peg")]:
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 2, w.get(key))
        r += 1
    poc = w.get("proof_of_cash", {}) or {}
    r += 1
    ws.cell(r, 1, "Proof of cash").font = BOLD
    r += 1
    for label, key in [("Reported revenue", "reported_revenue"), ("Cash receipts", "cash_receipts"), ("Variance", "variance"), ("Note", "note")]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, poc.get(key))
        r += 1


def _redflags(ws, rf):
    ws.cell(1, 1, "Red Flags").font = TITLE
    r = 3
    for j, h in enumerate(["Severity", "Category", "Title", "Rationale", "Recommendation", "Evidence source"]):
        ws.cell(r, 1 + j, h).font = BOLD
    r += 1
    order = {"high": 0, "medium": 1, "low": 2}
    for f in sorted(rf.get("findings", []), key=lambda x: order.get(x.get("severity"), 3)):
        ws.cell(r, 1, f.get("severity"))
        ws.cell(r, 2, f.get("category"))
        ws.cell(r, 3, f.get("title"))
        ws.cell(r, 4, f.get("rationale"))
        ws.cell(r, 5, f.get("recommendation"))
        ws.cell(r, 6, f.get("evidence_source"))
        r += 1


def _readiness(ws, rd):
    ws.cell(1, 1, f"Data-Room Readiness — {rd['score']}% ({rd['present_required']}/{rd['total_required']} required)").font = TITLE
    r = 3
    for j, h in enumerate(["Category", "Required", "Present", "Documents"]):
        ws.cell(r, 1 + j, h).font = BOLD
    r += 1
    for c in rd["categories"]:
        ws.cell(r, 1, c["category"])
        ws.cell(r, 2, "Yes" if c["required"] else "No")
        ws.cell(r, 3, "Yes" if c["present"] else "No")
        ws.cell(r, 4, ", ".join(c["documents"]))
        r += 1


def _profile(ws, profs):
    ws.cell(1, 1, "Data Profile").font = TITLE
    r = 3
    for p in profs:
        ws.cell(r, 1, f"{p['filename']}  ({p['n_rows']} rows x {p['n_cols']} cols)").font = BOLD
        r += 1
        for j, h in enumerate(["Column", "Type", "Fill %", "Distinct", "Summary"]):
            ws.cell(r, 1 + j, h).font = BOLD
        r += 1
        for col in p["columns"]:
            if col["type"] == "numeric":
                summary = f"min {col.get('min')} / max {col.get('max')} / mean {col.get('mean')} / sum {col.get('sum')}"
            else:
                summary = ", ".join(t["value"] for t in col.get("top", []))
            ws.cell(r, 1, col["name"])
            ws.cell(r, 2, col["type"])
            ws.cell(r, 3, round(100 - col["null_pct"], 1))
            ws.cell(r, 4, col["distinct"])
            ws.cell(r, 5, summary)
            r += 1
        r += 1


def _autosize(ws, cap: int = 60):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(cap, max(10, length + 2))
