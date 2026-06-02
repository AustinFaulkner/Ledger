"""Tabular data profiling — per-column statistics and data-quality flags for the
spreadsheet/CSV documents in a data room (a pandas-profiling / Great-Expectations-lite,
no pandas dependency).

For each table: row/column counts, and per column the inferred type, fill rate, distinct
count, numeric summary (min/max/mean/sum) or top categorical values. Surfaces the data-
quality issues an analyst checks first: missing values, constant columns, duplicate rows.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from . import store

TABULAR_KINDS = {"csv", "xlsx", "xls"}
_MAX_ROWS = 100_000  # guardrail for very large files


def profile_project(project_id: str) -> list[dict]:
    """Profile every tabular document in the project."""
    out = []
    for doc in store.list_documents(project_id):
        if (doc.get("kind") or "").lower() in TABULAR_KINDS:
            prof = profile_document(doc)
            if prof:
                out.append(prof)
    return out


def profile_document(doc: dict) -> dict | None:
    rows = _read_rows(doc)
    # Skip leading comment/title/blank lines (the sample CSVs open with a "# ..." title
    # before the real header row) so column names are detected correctly.
    while rows and (not any(str(c).strip() for c in rows[0]) or str(rows[0][0]).lstrip().startswith("#")):
        rows = rows[1:]
    if not rows or len(rows) < 1:
        return None
    header = [str(h).strip() or f"col{i+1}" for i, h in enumerate(rows[0])]
    data = rows[1:]
    ncols = len(header)

    columns = []
    for ci in range(ncols):
        values = [(r[ci] if ci < len(r) else "") for r in data]
        columns.append(_profile_column(header[ci], values))

    # table-level data-quality flags
    seen: set[tuple] = set()
    dup_rows = 0
    for r in data:
        key = tuple((r[ci] if ci < len(r) else "") for ci in range(ncols))
        if key in seen:
            dup_rows += 1
        else:
            seen.add(key)

    issues = []
    for c in columns:
        if c["null_pct"] >= 1:
            issues.append(f"{c['name']}: {c['null_pct']}% missing")
        if c["distinct"] <= 1 and c["count"] > 0:
            issues.append(f"{c['name']}: constant value")
    if dup_rows:
        issues.append(f"{dup_rows} duplicate row{'s' if dup_rows != 1 else ''}")

    return {
        "filename": doc["filename"],
        "kind": doc.get("kind", ""),
        "n_rows": len(data),
        "n_cols": ncols,
        "duplicate_rows": dup_rows,
        "columns": columns,
        "issues": issues,
    }


def _profile_column(name: str, values: list[str]) -> dict:
    non_empty = [v for v in values if str(v).strip() != ""]
    total = len(values)
    nulls = total - len(non_empty)
    distinct = len(set(non_empty))

    numbers = [n for n in (_to_number(v) for v in non_empty) if n is not None]
    is_numeric = bool(non_empty) and len(numbers) >= 0.8 * len(non_empty)

    col = {
        "name": name,
        "type": "numeric" if is_numeric else ("boolean" if _is_bool(non_empty) else "text"),
        "count": len(non_empty),
        "null_pct": round(100 * nulls / total, 1) if total else 0.0,
        "distinct": distinct,
    }
    if is_numeric and numbers:
        col["min"] = round(min(numbers), 2)
        col["max"] = round(max(numbers), 2)
        col["mean"] = round(sum(numbers) / len(numbers), 2)
        col["sum"] = round(sum(numbers), 2)
    else:
        freq: dict[str, int] = {}
        for v in non_empty:
            freq[str(v)] = freq.get(str(v), 0) + 1
        col["top"] = [
            {"value": k, "count": n}
            for k, n in sorted(freq.items(), key=lambda x: -x[1])[:3]
        ]
    return col


def _is_bool(values: list[str]) -> bool:
    if not values:
        return False
    s = {str(v).strip().lower() for v in values}
    return s <= {"true", "false", "yes", "no", "y", "n", "0", "1"} and len(s) <= 2


def _to_number(v: str) -> float | None:
    """Parse a financial-style cell: strips $ , %, treats (1,234) as negative."""
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return None


def _read_rows(doc: dict) -> list[list[str]]:
    path = Path(doc.get("path", ""))
    kind = (doc.get("kind") or "").lower()
    if not path.exists():
        return []
    if kind == "csv":
        text = path.read_text(encoding="utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        return [row for _, row in zip(range(_MAX_ROWS + 1), reader)]
    if kind in {"xlsx", "xls"}:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > _MAX_ROWS:
                    break
                rows.append(["" if c is None else str(c) for c in row])
            wb.close()
            return rows
        except Exception:
            return []
    return []
